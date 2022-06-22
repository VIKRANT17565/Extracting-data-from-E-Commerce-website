from data_collectoin import Data_Extraction
from openpyxl import Workbook
from openpyxl.styles import numbers

class Save_To_Excel(Data_Extraction):
    data_obj = Data_Extraction()
    # file = "testFile1.xlsx"
    book = Workbook()
    book.remove(book.active)

    # custom number format for INR Currency
    inr_format = "₹* #,##0.00;[Red]₹* -#,##0.00;₹* 0.00;* @"

    page_count = 10
    products_list = ["Smart phones"]
    
    # if you want to add random product in the <products_list>
    # then change the <get_data(product,page_count)> to <get_data_2(product,page_count)>
    for product in products_list:
        sheet = book.create_sheet(product)
        data = data_obj.get_data(product,page_count)
        
        
        for i in range(len(data)):
            sheet.append(data[i])
            sheet.cell(i+1, 2).number_format = inr_format

    data_obj.close_browser()
    book.save("DataBook.xlsx")